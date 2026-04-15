import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO

st.set_page_config(
    page_title="GROVE Procurement Hub",
    page_icon="🛒",
    layout="wide",
    initial_sidebar_state="expanded",
)

import db
from search import search_web, search_marketplace

# ── Constants ─────────────────────────────────────────
ALLOWED_DOMAINS = ["srkel.id", "teamup.id"]
CATEGORIES = [
    "Signage & Printing",
    "Event Supplies",
    "Office & Maintenance",
    "Jasa / Services",
]

# ── Session defaults ──────────────────────────────────
for key, default in [
    ("authenticated", False),
    ("user", None),
    ("cart", []),
    ("page", "Dashboard"),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ── Custom CSS ────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        font-size: 1.6rem; font-weight: 700; color: #1B2A4A;
        margin-bottom: 0.25rem;
    }
    .sub-header {
        font-size: 0.95rem; color: #666; margin-bottom: 1.5rem;
    }
    .stat-card {
        background: linear-gradient(135deg, #0F6E56 0%, #1B2A4A 100%);
        padding: 1.2rem; border-radius: 12px; color: white; text-align: center;
    }
    .stat-card h2 { font-size: 2rem; margin: 0; }
    .stat-card p { font-size: 0.85rem; margin: 0; opacity: 0.85; }
    .cart-badge {
        background: #D85A30; color: white; padding: 2px 8px;
        border-radius: 12px; font-size: 0.75rem; font-weight: 600;
    }
    div[data-testid="stSidebar"] { background-color: #1B2A4A; }
    div[data-testid="stSidebar"] .stMarkdown p,
    div[data-testid="stSidebar"] .stMarkdown h1,
    div[data-testid="stSidebar"] .stMarkdown h2,
    div[data-testid="stSidebar"] .stMarkdown h3,
    div[data-testid="stSidebar"] label { color: #E8E8E8 !important; }
    div[data-testid="stSidebar"] .stSelectbox label { color: #C5A355 !important; }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════════════
def check_domain(email):
    domain = email.split("@")[-1] if "@" in email else ""
    return domain in ALLOWED_DOMAINS

def login_page():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("---")
        st.markdown('<p class="main-header">🛒 GROVE Procurement Hub</p>', unsafe_allow_html=True)
        st.markdown('<p class="sub-header">Sistem Pengadaan Terpadu — GROVE at CIBIS</p>', unsafe_allow_html=True)

        has_any = db.has_users()

        if not has_any:
            st.info("🔧 Setup awal — buat akun Super Admin pertama.")
            with st.form("setup_form"):
                email = st.text_input("Email (@srkel.id / @teamup.id)")
                name = st.text_input("Nama Lengkap")
                pw = st.text_input("Password", type="password")
                pw2 = st.text_input("Konfirmasi Password", type="password")
                submit = st.form_submit_button("Buat Akun Admin", use_container_width=True)
            if submit:
                if not check_domain(email):
                    st.error("Domain email tidak diizinkan.")
                elif pw != pw2:
                    st.error("Password tidak cocok.")
                elif len(pw) < 6:
                    st.error("Password minimal 6 karakter.")
                else:
                    db.create_user(email.lower().strip(), name.strip(), "Admin", pw)
                    st.success("Akun Admin berhasil dibuat! Silakan login.")
                    st.rerun()
        else:
            with st.form("login_form"):
                email = st.text_input("Email")
                pw = st.text_input("Password", type="password")
                login = st.form_submit_button("Login", use_container_width=True)
            if login:
                if not check_domain(email):
                    st.error("Domain email tidak diizinkan.")
                    return
                user = db.verify_password(email.lower().strip(), pw)
                if user:
                    st.session_state["authenticated"] = True
                    st.session_state["user"] = user
                    st.rerun()
                else:
                    st.error("Email atau password salah.")


# ═══════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════
def sidebar():
    user = st.session_state["user"]
    is_admin = user.get("role") == "Admin"
    cart_count = len(st.session_state["cart"])

    with st.sidebar:
        st.markdown("### 🛒 GROVE Procurement Hub")
        st.markdown(f"👤 **{user.get('name', 'User')}** ({user.get('role', '')})")
        st.markdown("---")

        menu_items = ["Dashboard", "Vendor Directory", "Item Catalog", "Procurement Calculator", "Internet Digging"]
        if cart_count > 0:
            menu_items.append(f"🛒 Cart ({cart_count})")
        else:
            menu_items.append("🛒 Cart")

        if is_admin:
            menu_items.extend(["---", "⚙ Kelola Vendor", "⚙ Kelola Item", "⚙ Kelola User"])

        clean_items = [m for m in menu_items if m != "---"]
        divider_after = []
        if is_admin:
            for i, m in enumerate(menu_items):
                if m == "---":
                    divider_after.append(i)

        choice = st.radio("Navigasi", clean_items, label_visibility="collapsed")

        st.markdown("---")
        if st.button("🚪 Logout", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

    page = choice.replace(f" ({cart_count})", "") if cart_count > 0 else choice
    st.session_state["page"] = page


# ═══════════════════════════════════════════════════════
# PAGES
# ═══════════════════════════════════════════════════════

def page_dashboard():
    st.markdown('<p class="main-header">📊 Dashboard</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Ringkasan data procurement GROVE at CIBIS</p>', unsafe_allow_html=True)

    stats = db.get_dashboard_stats()
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f"""<div class="stat-card"><h2>{stats['active_vendors']}</h2><p>Vendor Aktif</p></div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class="stat-card"><h2>{stats['total_items']}</h2><p>Total Item</p></div>""", unsafe_allow_html=True)
    with c3:
        st.markdown(f"""<div class="stat-card"><h2>{stats['category_count']}</h2><p>Kategori</p></div>""", unsafe_allow_html=True)
    with c4:
        cart_n = len(st.session_state["cart"])
        st.markdown(f"""<div class="stat-card"><h2>{cart_n}</h2><p>Item di Cart</p></div>""", unsafe_allow_html=True)

    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("##### Vendor per Kategori")
        vendors = db.get_vendors()
        if vendors:
            cat_counts = {}
            for v in vendors:
                c = v.get("category", "Lainnya")
                cat_counts[c] = cat_counts.get(c, 0) + 1
            df_cat = pd.DataFrame(list(cat_counts.items()), columns=["Kategori", "Jumlah"])
            st.bar_chart(df_cat.set_index("Kategori"))
        else:
            st.info("Belum ada data vendor.")

    with col2:
        st.markdown("##### Item per Kategori")
        items = db.get_items()
        if items:
            cat_counts = {}
            for it in items:
                c = it.get("category", "Lainnya")
                cat_counts[c] = cat_counts.get(c, 0) + 1
            df_cat = pd.DataFrame(list(cat_counts.items()), columns=["Kategori", "Jumlah"])
            st.bar_chart(df_cat.set_index("Kategori"))
        else:
            st.info("Belum ada data item.")


def page_vendor_directory():
    st.markdown('<p class="main-header">🏢 Vendor Directory</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Cari dan jelajahi vendor yang terdaftar</p>', unsafe_allow_html=True)

    col1, col2 = st.columns([2, 1])
    with col1:
        search_q = st.text_input("🔍 Cari vendor", placeholder="Nama, kategori, atau keyword...")
    with col2:
        cat_filter = st.selectbox("Kategori", ["Semua"] + CATEGORIES)

    vendors = db.get_vendors()
    if search_q:
        q = search_q.lower()
        vendors = [v for v in vendors if q in str(v).lower()]
    if cat_filter != "Semua":
        vendors = [v for v in vendors if v.get("category") == cat_filter]

    if not vendors:
        st.info("Tidak ada vendor ditemukan. Coba ubah filter atau gunakan **Internet Digging**.")
        return

    st.markdown(f"**{len(vendors)} vendor ditemukan**")
    for v in vendors:
        status_icon = "🟢" if v.get("status") == "Active" else "🔴"
        rating = v.get("rating", 0)
        stars = "⭐" * int(float(rating)) if rating else ""
        with st.expander(f"{status_icon} **{v.get('vendor_name', 'N/A')}** — {v.get('category', '')} {stars}"):
            c1, c2 = st.columns(2)
            with c1:
                st.write(f"**ID:** {v.get('vendor_id', '')}")
                st.write(f"**Sub-kategori:** {v.get('subcategory', '-')}")
                st.write(f"**PIC:** {v.get('contact_person', '-')}")
                st.write(f"**Telp:** {v.get('phone', '-')}")
            with c2:
                st.write(f"**Email:** {v.get('email', '-')}")
                st.write(f"**Alamat:** {v.get('address', '-')}")
                st.write(f"**Rating:** {rating}/5 {stars}")
                st.write(f"**Catatan:** {v.get('notes', '-')}")

            items = db.get_items()
            vendor_items = [i for i in items if i.get("vendor_id") == v.get("vendor_id")]
            if vendor_items:
                st.markdown("**Item yang tersedia:**")
                df = pd.DataFrame(vendor_items)[["item_id", "item_name", "unit", "unit_price", "lead_time_days"]]
                df.columns = ["ID", "Item", "Satuan", "Harga (IDR)", "Lead Time (hari)"]
                df["Harga (IDR)"] = df["Harga (IDR)"].apply(lambda x: f"Rp {int(x):,}" if x else "-")
                st.dataframe(df, use_container_width=True, hide_index=True)


def page_item_catalog():
    st.markdown('<p class="main-header">📦 Item Catalog</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Browse semua item dan harga satuan</p>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        search_q = st.text_input("🔍 Cari item", placeholder="Nama item, vendor, atau keyword...")
    with col2:
        cat_filter = st.selectbox("Kategori", ["Semua"] + CATEGORIES, key="item_cat")
    with col3:
        sort_by = st.selectbox("Urutkan", ["Nama A-Z", "Harga Terendah", "Harga Tertinggi"])

    items = db.get_items()
    vendors = {v["vendor_id"]: v["vendor_name"] for v in db.get_vendors()}

    if search_q:
        q = search_q.lower()
        items = [i for i in items if q in str(i).lower()]
    if cat_filter != "Semua":
        items = [i for i in items if i.get("category") == cat_filter]

    if sort_by == "Harga Terendah":
        items.sort(key=lambda x: float(x.get("unit_price", 0)))
    elif sort_by == "Harga Tertinggi":
        items.sort(key=lambda x: float(x.get("unit_price", 0)), reverse=True)
    else:
        items.sort(key=lambda x: x.get("item_name", ""))

    if not items:
        st.info("Tidak ada item ditemukan. Coba **Internet Digging** untuk mencari di web.")
        return

    st.markdown(f"**{len(items)} item ditemukan**")

    for it in items:
        price = int(float(it.get("unit_price", 0)))
        vendor_name = vendors.get(it.get("vendor_id", ""), "Unknown")
        col1, col2, col3 = st.columns([3, 1, 1])
        with col1:
            st.markdown(f"**{it.get('item_name', '')}** — {it.get('description', '')}")
            st.caption(f"Vendor: {vendor_name} | Satuan: {it.get('unit', '')} | Lead: {it.get('lead_time_days', 0)} hari")
        with col2:
            st.markdown(f"### Rp {price:,}")
            st.caption(f"per {it.get('unit', 'pcs')}")
        with col3:
            qty = st.number_input("QTY", min_value=0, value=0, step=1, key=f"qty_{it.get('item_id', '')}")
            if qty > 0:
                if st.button("➕ Cart", key=f"add_{it.get('item_id', '')}"):
                    st.session_state["cart"].append({
                        "item_id": it.get("item_id", ""),
                        "vendor_id": it.get("vendor_id", ""),
                        "item_name": it.get("item_name", ""),
                        "vendor_name": vendor_name,
                        "unit": it.get("unit", ""),
                        "unit_price": price,
                        "qty": qty,
                    })
                    st.success(f"✅ {it['item_name']} x{qty} ditambahkan!")
                    st.rerun()
        st.markdown("---")


def page_calculator():
    st.markdown('<p class="main-header">🧮 Procurement Calculator</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Pilih item, masukkan QTY, hitung total biaya</p>', unsafe_allow_html=True)

    items = db.get_items()
    vendors_map = {v["vendor_id"]: v["vendor_name"] for v in db.get_vendors()}

    if not items:
        st.info("Belum ada item di database. Tambahkan melalui menu Kelola Item.")
        return

    cat_filter = st.selectbox("Filter Kategori", ["Semua"] + CATEGORIES, key="calc_cat")
    filtered = items if cat_filter == "Semua" else [i for i in items if i.get("category") == cat_filter]

    item_options = {f"{i['item_name']} — {vendors_map.get(i['vendor_id'], '?')} (Rp {int(float(i.get('unit_price', 0))):,}/{i.get('unit', 'pcs')})": i for i in filtered}

    selected_labels = st.multiselect("Pilih Item", list(item_options.keys()))

    if selected_labels:
        st.markdown("##### Detail Kalkulasi")
        calc_rows = []
        total = 0
        for label in selected_labels:
            it = item_options[label]
            price = int(float(it.get("unit_price", 0)))
            qty = st.number_input(
                f"QTY — {it['item_name']}",
                min_value=int(float(it.get("min_order", 1))),
                value=int(float(it.get("min_order", 1))),
                step=1,
                key=f"calc_{it['item_id']}",
            )
            subtotal = qty * price
            total += subtotal
            calc_rows.append({
                "Item": it["item_name"],
                "Vendor": vendors_map.get(it["vendor_id"], "?"),
                "Satuan": it.get("unit", ""),
                "Harga": f"Rp {price:,}",
                "QTY": qty,
                "Subtotal": f"Rp {subtotal:,}",
                "_item": it,
                "_qty": qty,
                "_price": price,
            })

        if calc_rows:
            df = pd.DataFrame(calc_rows)[["Item", "Vendor", "Satuan", "Harga", "QTY", "Subtotal"]]
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.markdown(f"### 💰 Grand Total: **Rp {total:,}**")

            if st.button("🛒 Tambahkan semua ke Cart", use_container_width=True):
                for row in calc_rows:
                    it = row["_item"]
                    st.session_state["cart"].append({
                        "item_id": it.get("item_id", ""),
                        "vendor_id": it.get("vendor_id", ""),
                        "item_name": it.get("item_name", ""),
                        "vendor_name": vendors_map.get(it["vendor_id"], "?"),
                        "unit": it.get("unit", ""),
                        "unit_price": row["_price"],
                        "qty": row["_qty"],
                    })
                st.success(f"✅ {len(calc_rows)} item ditambahkan ke cart!")
                st.rerun()


def page_internet_digging():
    st.markdown('<p class="main-header">🌐 Internet Digging</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Cari vendor & harga di internet ketika item tidak ada di database</p>', unsafe_allow_html=True)

    col1, col2 = st.columns([3, 1])
    with col1:
        query = st.text_input("🔍 Apa yang Anda cari?", placeholder="contoh: backdrop event outdoor 3x2 meter")
    with col2:
        search_type = st.selectbox("Sumber", ["Web Umum", "Marketplace (Tokopedia/Shopee)"])

    if st.button("🚀 Cari di Internet", use_container_width=True, disabled=not query):
        user_email = st.session_state["user"].get("email", "")
        with st.spinner("Mencari di internet..."):
            if search_type == "Marketplace (Tokopedia/Shopee)":
                results = search_marketplace(query)
            else:
                results = search_web(query)
            db.log_search(user_email, query, search_type, len(results))

        if not results:
            st.warning("Tidak ditemukan hasil. Coba keyword yang berbeda.")
            return

        st.success(f"**{len(results)} hasil ditemukan**")

        for idx, r in enumerate(results):
            with st.expander(f"🔗 {r['title'][:80]}"):
                st.write(r.get("body", ""))
                st.caption(r.get("url", ""))

                is_admin = st.session_state["user"].get("role") == "Admin"
                if is_admin:
                    st.markdown("---")
                    st.markdown("**💾 Simpan ke Database (Admin)**")
                    with st.form(f"save_form_{idx}"):
                        c1, c2 = st.columns(2)
                        with c1:
                            v_name = st.text_input("Nama Vendor", key=f"sv_{idx}")
                            v_cat = st.selectbox("Kategori", CATEGORIES, key=f"sc_{idx}")
                        with c2:
                            i_name = st.text_input("Nama Item", key=f"si_{idx}")
                            i_price = st.number_input("Harga Satuan (IDR)", min_value=0, step=1000, key=f"sp_{idx}")
                        i_unit = st.text_input("Satuan", value="pcs", key=f"su_{idx}")
                        save_btn = st.form_submit_button("💾 Simpan ke Database")

                    if save_btn and v_name and i_name:
                        vendor_id = db.add_vendor({
                            "vendor_name": v_name,
                            "category": v_cat,
                            "subcategory": "",
                            "notes": f"Source: Internet Digging — {r.get('url', '')}",
                        })
                        db.add_item({
                            "vendor_id": vendor_id,
                            "category": v_cat,
                            "item_name": i_name,
                            "unit": i_unit,
                            "unit_price": i_price,
                            "source": "Internet",
                        })
                        st.success(f"✅ Vendor '{v_name}' & Item '{i_name}' tersimpan!")
                        st.rerun()


def page_cart():
    st.markdown('<p class="main-header">🛒 Procurement Cart</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Review dan export daftar belanja</p>', unsafe_allow_html=True)

    cart = st.session_state["cart"]
    if not cart:
        st.info("Cart kosong. Tambahkan item dari **Item Catalog**, **Calculator**, atau **Internet Digging**.")
        return

    rows = []
    total = 0
    for i, item in enumerate(cart):
        subtotal = item["qty"] * item["unit_price"]
        total += subtotal
        rows.append({
            "#": i + 1,
            "Item": item["item_name"],
            "Vendor": item.get("vendor_name", "-"),
            "Satuan": item.get("unit", ""),
            "Harga": f"Rp {item['unit_price']:,}",
            "QTY": item["qty"],
            "Subtotal": f"Rp {subtotal:,}",
        })

    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)

    st.markdown(f"### 💰 Grand Total: **Rp {total:,}**")
    st.caption(f"{len(cart)} item dari {len(set(i.get('vendor_name', '') for i in cart))} vendor")

    st.markdown("---")
    col1, col2, col3 = st.columns(3)

    with col1:
        purpose = st.text_input("Tujuan Pengadaan", placeholder="contoh: Event Natal 2026")

    with col2:
        if st.button("💾 Simpan ke Database", use_container_width=True, disabled=not purpose):
            user_email = st.session_state["user"].get("email", "")
            cart_id = db.save_cart(cart, user_email, purpose)
            st.success(f"✅ Cart disimpan! ID: {cart_id}")

    with col3:
        if st.button("📥 Export Excel", use_container_width=True):
            excel_data = export_cart_excel(cart, total, purpose)
            st.download_button(
                label="⬇ Download Excel",
                data=excel_data,
                file_name=f"GROVE_Procurement_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    st.markdown("---")
    st.markdown("##### Kelola Cart")
    remove_idx = st.number_input("Hapus item ke-", min_value=1, max_value=len(cart), value=1, step=1)
    if st.button("🗑 Hapus item"):
        st.session_state["cart"].pop(remove_idx - 1)
        st.rerun()

    if st.button("🗑 Kosongkan Cart", type="secondary"):
        st.session_state["cart"] = []
        st.rerun()


def export_cart_excel(cart, total, purpose):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Procurement Summary"

    navy = "1B2A4A"
    teal = "0F6E56"
    gold = "C5A355"
    hfont = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill("solid", fgColor=navy)
    tfont = Font(name="Arial", size=10)
    bdr = Border(
        left=Side("thin", color="D9D9D9"), right=Side("thin", color="D9D9D9"),
        top=Side("thin", color="D9D9D9"), bottom=Side("thin", color="D9D9D9"),
    )

    ws.merge_cells("A1:G1")
    ws["A1"] = "GROVE at CIBIS — Procurement Summary"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=navy)

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Tanggal: {datetime.now().strftime('%d %B %Y')} | Tujuan: {purpose or '-'}"
    ws["A2"].font = Font(name="Arial", size=10, color="666666")

    headers = ["No", "Item", "Vendor", "Satuan", "Harga Satuan", "QTY", "Subtotal"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr

    for i, item in enumerate(cart):
        r = 5 + i
        subtotal = item["qty"] * item["unit_price"]
        vals = [i + 1, item["item_name"], item.get("vendor_name", "-"), item.get("unit", ""), item["unit_price"], item["qty"], subtotal]
        alt = PatternFill("solid", fgColor="F2F2F2") if i % 2 else PatternFill("solid", fgColor="FFFFFF")
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = tfont
            cell.fill = alt
            cell.border = bdr
        ws.cell(row=r, column=5).number_format = '#,##0'
        ws.cell(row=r, column=7).number_format = '#,##0'

    total_row = 5 + len(cart)
    ws.merge_cells(f"A{total_row}:F{total_row}")
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=teal)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=7, value=total).font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws.cell(row=total_row, column=7).fill = PatternFill("solid", fgColor=teal)
    ws.cell(row=total_row, column=7).number_format = '#,##0'

    sig_row = total_row + 3
    ws.cell(row=sig_row, column=1, value="Dibuat oleh:").font = Font(name="Arial", size=10, italic=True)
    ws.cell(row=sig_row, column=4, value="Disetujui oleh:").font = Font(name="Arial", size=10, italic=True)
    ws.cell(row=sig_row + 3, column=1, value="(___________________)").font = tfont
    ws.cell(row=sig_row + 3, column=4, value="(___________________)").font = tfont

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════
# ADMIN PAGES
# ═══════════════════════════════════════════════════════

def page_admin_vendors():
    st.markdown('<p class="main-header">⚙ Kelola Vendor</p>', unsafe_allow_html=True)

    tab_add, tab_edit = st.tabs(["➕ Tambah Vendor", "✏️ Edit Vendor"])

    with tab_add:
        with st.form("add_vendor_form"):
            c1, c2 = st.columns(2)
            with c1:
                v_name = st.text_input("Nama Vendor *")
                v_cat = st.selectbox("Kategori *", CATEGORIES)
                v_subcat = st.text_input("Sub-kategori")
                v_contact = st.text_input("Contact Person")
            with c2:
                v_phone = st.text_input("No. Telp")
                v_email = st.text_input("Email")
                v_address = st.text_area("Alamat", height=68)
                v_rating = st.slider("Rating", 1.0, 5.0, 4.0, 0.5)
            v_notes = st.text_area("Catatan", height=68)
            submit = st.form_submit_button("💾 Simpan Vendor", use_container_width=True)

        if submit:
            if not v_name:
                st.error("Nama vendor wajib diisi.")
            else:
                vid = db.add_vendor({
                    "vendor_name": v_name, "category": v_cat, "subcategory": v_subcat,
                    "contact_person": v_contact, "phone": v_phone, "email": v_email,
                    "address": v_address, "rating": v_rating, "notes": v_notes,
                })
                st.success(f"✅ Vendor '{v_name}' disimpan dengan ID: {vid}")

    with tab_edit:
        vendors = db.get_vendors()
        if not vendors:
            st.info("Belum ada vendor.")
            return
        vendor_map = {f"{v['vendor_id']} — {v['vendor_name']}": v for v in vendors}
        selected = st.selectbox("Pilih Vendor", list(vendor_map.keys()))
        v = vendor_map[selected]

        with st.form("edit_vendor_form"):
            c1, c2 = st.columns(2)
            with c1:
                e_name = st.text_input("Nama Vendor", value=v.get("vendor_name", ""))
                e_cat = st.selectbox("Kategori", CATEGORIES, index=CATEGORIES.index(v["category"]) if v.get("category") in CATEGORIES else 0)
                e_contact = st.text_input("Contact Person", value=v.get("contact_person", ""))
                e_phone = st.text_input("No. Telp", value=v.get("phone", ""))
            with c2:
                e_email = st.text_input("Email", value=v.get("email", ""))
                e_address = st.text_area("Alamat", value=v.get("address", ""), height=68)
                e_rating = st.slider("Rating", 1.0, 5.0, float(v.get("rating", 4.0)), 0.5)
                e_status = st.selectbox("Status", ["Active", "Inactive"], index=0 if v.get("status") == "Active" else 1)
            e_notes = st.text_area("Catatan", value=v.get("notes", ""), height=68)
            update = st.form_submit_button("💾 Update Vendor", use_container_width=True)

        if update:
            db.update_vendor(v["vendor_id"], {
                "vendor_name": e_name, "category": e_cat, "contact_person": e_contact,
                "phone": e_phone, "email": e_email, "address": e_address,
                "rating": e_rating, "notes": e_notes, "status": e_status,
            })
            st.success(f"✅ Vendor '{e_name}' berhasil diupdate!")
            st.rerun()


def page_admin_items():
    st.markdown('<p class="main-header">⚙ Kelola Item</p>', unsafe_allow_html=True)

    vendors = db.get_vendors()
    vendor_options = {f"{v['vendor_id']} — {v['vendor_name']}": v["vendor_id"] for v in vendors if v.get("status") == "Active"}

    if not vendor_options:
        st.warning("Belum ada vendor aktif. Tambahkan vendor terlebih dahulu.")
        return

    tab_add, tab_edit = st.tabs(["➕ Tambah Item", "✏️ Edit Item"])

    with tab_add:
        with st.form("add_item_form"):
            c1, c2 = st.columns(2)
            with c1:
                i_vendor = st.selectbox("Vendor *", list(vendor_options.keys()))
                i_cat = st.selectbox("Kategori *", CATEGORIES)
                i_name = st.text_input("Nama Item *")
                i_desc = st.text_input("Deskripsi")
            with c2:
                i_unit = st.text_input("Satuan", value="pcs")
                i_price = st.number_input("Harga Satuan (IDR) *", min_value=0, step=1000)
                i_min = st.number_input("Min. Order", min_value=1, value=1)
                i_lead = st.number_input("Lead Time (hari)", min_value=0, value=1)
            submit = st.form_submit_button("💾 Simpan Item", use_container_width=True)

        if submit:
            if not i_name or not i_price:
                st.error("Nama item dan harga wajib diisi.")
            else:
                iid = db.add_item({
                    "vendor_id": vendor_options[i_vendor],
                    "category": i_cat, "item_name": i_name,
                    "description": i_desc, "unit": i_unit,
                    "unit_price": i_price, "min_order": i_min,
                    "lead_time_days": i_lead, "source": "Manual",
                })
                st.success(f"✅ Item '{i_name}' disimpan dengan ID: {iid}")

    with tab_edit:
        items = db.get_items()
        if not items:
            st.info("Belum ada item.")
            return
        item_map = {f"{i['item_id']} — {i['item_name']}": i for i in items}
        selected = st.selectbox("Pilih Item", list(item_map.keys()))
        it = item_map[selected]

        with st.form("edit_item_form"):
            c1, c2 = st.columns(2)
            with c1:
                e_name = st.text_input("Nama Item", value=it.get("item_name", ""))
                e_desc = st.text_input("Deskripsi", value=it.get("description", ""))
                current_vid = it.get("vendor_id", "")
                vendor_keys = list(vendor_options.keys())
                default_idx = next((i for i, k in enumerate(vendor_keys) if current_vid in k), 0)
                e_vendor = st.selectbox("Vendor", vendor_keys, index=default_idx)
            with c2:
                e_unit = st.text_input("Satuan", value=it.get("unit", "pcs"))
                e_price = st.number_input("Harga Satuan (IDR)", value=int(float(it.get("unit_price", 0))), step=1000)
                e_min = st.number_input("Min. Order", value=int(float(it.get("min_order", 1))))
                e_lead = st.number_input("Lead Time (hari)", value=int(float(it.get("lead_time_days", 0))))
            update = st.form_submit_button("💾 Update Item", use_container_width=True)

        if update:
            db.update_item(it["item_id"], {
                "vendor_id": vendor_options[e_vendor],
                "item_name": e_name, "description": e_desc,
                "unit": e_unit, "unit_price": e_price,
                "min_order": e_min, "lead_time_days": e_lead,
            })
            st.success(f"✅ Item '{e_name}' berhasil diupdate!")
            st.rerun()


def page_admin_users():
    st.markdown('<p class="main-header">⚙ Kelola User</p>', unsafe_allow_html=True)

    users = db.get_all_users()
    if users:
        df = pd.DataFrame(users)[["email", "name", "role", "created_at"]]
        df.columns = ["Email", "Nama", "Role", "Dibuat"]
        st.dataframe(df, use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("##### Tambah User Baru")
    with st.form("add_user_form"):
        c1, c2, c3 = st.columns(3)
        with c1:
            u_email = st.text_input("Email (@srkel.id / @teamup.id)")
        with c2:
            u_name = st.text_input("Nama Lengkap")
        with c3:
            u_role = st.selectbox("Role", ["Staff", "Admin"])
        u_pw = st.text_input("Password", type="password")
        submit = st.form_submit_button("➕ Tambah User", use_container_width=True)

    if submit:
        if not check_domain(u_email):
            st.error("Domain email tidak diizinkan.")
        elif not u_name or not u_pw:
            st.error("Semua field wajib diisi.")
        elif len(u_pw) < 6:
            st.error("Password minimal 6 karakter.")
        else:
            existing = db.get_user(u_email.lower().strip())
            if existing:
                st.error("Email sudah terdaftar.")
            else:
                db.create_user(u_email.lower().strip(), u_name.strip(), u_role, u_pw)
                st.success(f"✅ User '{u_name}' berhasil ditambahkan!")
                st.rerun()


# ═══════════════════════════════════════════════════════
# MAIN ROUTER
# ═══════════════════════════════════════════════════════
def main():
    if not st.session_state["authenticated"]:
        login_page()
        return

    sidebar()
    page = st.session_state["page"]

    if page == "Dashboard":
        page_dashboard()
    elif page == "Vendor Directory":
        page_vendor_directory()
    elif page == "Item Catalog":
        page_item_catalog()
    elif page == "Procurement Calculator":
        page_calculator()
    elif page == "Internet Digging":
        page_internet_digging()
    elif page.startswith("🛒"):
        page_cart()
    elif page == "⚙ Kelola Vendor":
        page_admin_vendors()
    elif page == "⚙ Kelola Item":
        page_admin_items()
    elif page == "⚙ Kelola User":
        page_admin_users()
    else:
        page_dashboard()

if __name__ == "__main__":
    main()
